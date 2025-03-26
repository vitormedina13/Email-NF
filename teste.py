import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import io
import os
from pathlib import Path

st.set_page_config(page_title="Transferência de Dados de Câmbio", layout="wide")

st.title("Transferência de Dados de Câmbio")

# Função para ler os dados do arquivo de origem
def ler_dados_cambio(arquivo_cambio, data_inicial, data_final):
    try:
        # Usar pandas para ler os dados
        df = pd.read_excel(
            arquivo_cambio, 
            sheet_name="BGP e BGX Cambio",
            engine="openpyxl"
        )
        
        # Extrair as colunas necessárias
        dados = pd.DataFrame({
            "Data": df.iloc[:, 1],  # Coluna B (index 1)
            "Cliente": df.iloc[:, 19],  # Coluna T (index 19)
            "Receita_BGX": df.iloc[:, 47]  # Coluna AV (index 47)
        })
        
        # Remover linhas com data vazia
        dados = dados.dropna(subset=["Data"])
        
        # Converter a coluna Data para datetime
        dados["Data"] = pd.to_datetime(dados["Data"], errors='coerce')
        dados = dados.dropna(subset=["Data"])  # Remover linhas com datas inválidas
        
        # Filtrar por data
        if data_inicial and data_final:
            data_inicial = pd.to_datetime(data_inicial)
            data_final = pd.to_datetime(data_final)
            dados = dados[(dados["Data"] >= data_inicial) & (dados["Data"] <= data_final)]
        
        return dados
    
    except Exception as e:
        st.error(f"Erro ao ler o arquivo de câmbio: {str(e)}")
        return None

# Função para ler os dados do arquivo de destino
def ler_dados_destino(arquivo_nf):
    try:
        # Ler a aba específica
        df_destino = pd.read_excel(
            arquivo_nf,
            sheet_name="Todas as Op - Câmbio",
            engine="openpyxl"
        )
        
        # Pegar apenas as colunas relevantes para visualização
        colunas_relevantes = {
            0: "Data",           # Coluna A
            4: "Receita_BGX",    # Coluna E
            8: "Cliente"         # Coluna I
        }
        
        dados_visualizacao = pd.DataFrame()
        for idx, nome in colunas_relevantes.items():
            if idx < len(df_destino.columns):
                dados_visualizacao[nome] = df_destino.iloc[:, idx]
        
        return dados_visualizacao, df_destino
    
    except Exception as e:
        st.error(f"Erro ao ler o arquivo de destino: {str(e)}")
        return None, None

# Função para atualizar os dados - CORRIGIDA
def atualizar_dados(df_destino, dados_cambio):
    try:
        # Preparar os dados para inserção
        novos_dados = pd.DataFrame({
            "Data": dados_cambio["Data"],
            "Cliente": dados_cambio["Cliente"],
            "Receita_BGX": dados_cambio["Receita_BGX"]
        })
        
        # Imprimir para debug
        st.write("Novos dados a inserir:", novos_dados.shape)
        
        # Criar uma cópia do dataframe de destino
        df_atualizado = df_destino.copy()
        
        # Obter o índice da última linha não vazia (verificando apenas a primeira coluna)
        ultima_linha = df_destino.shape[0]
        for i in range(df_destino.shape[0]-1, -1, -1):
            if pd.notna(df_destino.iloc[i, 0]):  # Verifica se a primeira coluna não é NaN
                ultima_linha = i + 1
                break
        
        st.write(f"Última linha ocupada: {ultima_linha}")
        
        # Criar um novo DataFrame para armazenar os resultados
        # Isso garante que preservamos a estrutura original do DataFrame de destino
        novas_linhas = []
        
        # Para cada linha nos novos dados
        for _, row in novos_dados.iterrows():
            # Criar uma nova linha com a mesma estrutura do df_destino
            nova_linha = [None] * len(df_destino.columns)
            
            # Preencher apenas as colunas específicas
            nova_linha[0] = row["Data"]         # Coluna A (índice 0)
            nova_linha[4] = row["Receita_BGX"]  # Coluna E (índice 4)
            nova_linha[8] = row["Cliente"]      # Coluna I (índice 8)
            
            # Adicionar à lista de novas linhas
            novas_linhas.append(nova_linha)
        
        # Converter a lista de novas linhas para DataFrame
        novas_linhas_df = pd.DataFrame(novas_linhas, columns=df_destino.columns)
        
        # Adicionar as novas linhas ao DataFrame atualizado
        df_atualizado = pd.concat([df_atualizado, novas_linhas_df], ignore_index=True)
        
        st.write(f"Novas linhas adicionadas: {len(novas_linhas)}")
        st.write(f"Tamanho final do DataFrame: {df_atualizado.shape}")
        
        return df_atualizado
    except Exception as e:
        st.error(f"Erro ao atualizar os dados: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

# Função para salvar dados no Excel
def salvar_em_excel(df, arquivo_original=None):
    try:
        output = io.BytesIO()
        
        if arquivo_original is not None:
            # Tentar preservar a estrutura original
            try:
                # Ler o workbook original
                wb = openpyxl.load_workbook(arquivo_original, keep_vba=True)
                
                # Obter a planilha de destino
                sheet_name = "Todas as Op - Câmbio"
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Limpar a planilha (manter cabeçalho)
                    for row in list(ws.rows)[1:]:
                        for cell in row:
                            cell.value = None
                    
                    # Escrever os novos dados
                    for r_idx, row in df.iterrows():
                        for c_idx, value in enumerate(row):
                            ws.cell(row=r_idx+2, column=c_idx+1).value = value
                    
                    # Salvar no buffer
                    wb.save(output)
                else:
                    raise Exception(f"Aba {sheet_name} não encontrada")
            except Exception as e:
                st.warning(f"Não foi possível preservar a estrutura original. Criando novo arquivo. Erro: {str(e)}")
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Dados_Atualizados', index=False)
        else:
            # Criar um novo arquivo
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Dados_Atualizados', index=False)
        
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Erro ao salvar os dados: {str(e)}")
        return None

# Interface Streamlit
st.header("Selecione os Arquivos")

# Upload de arquivos
arquivo_cambio_upload = st.file_uploader("Selecione o arquivo de Operações de câmbio", type=["xlsm", "xlsx"])
arquivo_nf_upload = st.file_uploader("Selecione o arquivo de Notas Fiscais", type=["xlsm", "xlsx"])

# Seleção de datas
col1, col2 = st.columns(2)
with col1:
    data_inicial = st.date_input("Data Inicial", value=None)
with col2:
    data_final = st.date_input("Data Final", value=None)

# Variáveis para armazenar os dados
dados_destino = None
df_destino_completo = None
dados_cambio = None
df_atualizado = None

# Criar containers para armazenar o estado da aplicação
if 'dados_carregados' not in st.session_state:
    st.session_state.dados_carregados = False
    
if 'dados_destino_carregados' not in st.session_state:
    st.session_state.dados_destino_carregados = False
    
if 'df_destino_completo' not in st.session_state:
    st.session_state.df_destino_completo = None
    
if 'dados_cambio' not in st.session_state:
    st.session_state.dados_cambio = None

# Tab layout
tab1, tab2, tab3 = st.tabs(["Arquivo de Origem", "Arquivo de Destino", "Arquivo Atualizado"])

# Processar e exibir os dados quando os arquivos forem carregados
if arquivo_cambio_upload:
    with tab1:
        if st.button("Carregar Dados de Câmbio"):
            with st.spinner("Processando arquivo de origem..."):
                dados_cambio = ler_dados_cambio(arquivo_cambio_upload, data_inicial, data_final)
                
                if dados_cambio is not None and not dados_cambio.empty:
                    st.success(f"Dados de câmbio carregados com sucesso! {len(dados_cambio)} registros encontrados.")
                    st.session_state.dados_carregados = True
                    st.session_state.dados_cambio = dados_cambio
                    
                    # Mostrar os dados
                    st.subheader("Dados a serem transferidos:")
                    st.dataframe(dados_cambio)
                    
                    # Oferecer download dos dados extraídos
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        dados_cambio.to_excel(writer, sheet_name='Dados_Extraídos', index=False)
                    buffer.seek(0)
                    
                    st.download_button(
                        label="Baixar Dados Extraídos (XLSX)",
                        data=buffer,
                        file_name="Dados_Cambio_Extraidos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("Nenhum dado encontrado para o período selecionado.")

if arquivo_nf_upload:
    with tab2:
        if st.button("Carregar Arquivo de Destino"):
            with st.spinner("Processando arquivo de destino..."):
                dados_destino, df_destino_completo = ler_dados_destino(arquivo_nf_upload)
                
                if dados_destino is not None:
                    st.success("Arquivo de destino carregado com sucesso!")
                    st.session_state.dados_destino_carregados = True
                    st.session_state.df_destino_completo = df_destino_completo
                    
                    # Mostrar os dados
                    st.subheader("Dados atuais no arquivo de destino:")
                    st.dataframe(dados_destino)
                    
                    # Oferecer download do arquivo original
                    buffer = io.BytesIO()
                    arquivo_nf_upload.seek(0)
                    buffer.write(arquivo_nf_upload.read())
                    buffer.seek(0)
                    
                    st.download_button(
                        label="Baixar Arquivo de Destino Original",
                        data=buffer,
                        file_name="Arquivo_Destino_Original.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("Não foi possível carregar o arquivo de destino.")

# Botão para combinar os dados
with tab3:
    if (st.session_state.dados_carregados and 
        st.session_state.dados_destino_carregados and 
        st.session_state.dados_cambio is not None and 
        st.session_state.df_destino_completo is not None):
        
        if st.button("Combinar Dados"):
            with st.spinner("Combinando dados..."):
                df_atualizado = atualizar_dados(st.session_state.df_destino_completo, st.session_state.dados_cambio)
                
                if df_atualizado is not None:
                    st.success("Dados combinados com sucesso!")
                    
                    # Mostrar os dados atualizados (apenas as colunas relevantes)
                    st.subheader("Dados atualizados:")
                    colunas_visualizacao = {
                        0: "Data",
                        4: "Receita_BGX",
                        8: "Cliente"
                    }
                    
                    dados_visualizacao = pd.DataFrame()
                    for idx, nome in colunas_visualizacao.items():
                        if idx < len(df_atualizado.columns):
                            dados_visualizacao[nome] = df_atualizado.iloc[:, idx]
                    
                    # Mostrar uma amostra maior dos dados para debug
                    st.write("Amostra dos primeiros 20 registros:")
                    st.dataframe(dados_visualizacao.head(20))
                    
                    # Oferecer download do arquivo atualizado
                    buffer_novo = salvar_em_excel(df_atualizado, arquivo_nf_upload)
                    buffer_simples = salvar_em_excel(df_atualizado, None)
                    
                    if buffer_novo is not None:
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.download_button(
                                label="Baixar Arquivo Completo Atualizado",
                                data=buffer_novo,
                                file_name="Arquivo_Destino_Atualizado.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        with col2:
                            st.download_button(
                                label="Baixar Apenas os Dados (Formato Simples)",
                                data=buffer_simples,
                                file_name="Dados_Atualizados_Simples.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        # Adicionar opção de CSV
                        csv_buffer = io.StringIO()
                        dados_visualizacao.to_csv(csv_buffer, index=False)
                        
                        st.download_button(
                            label="Baixar Dados como CSV",
                            data=csv_buffer.getvalue(),
                            file_name="Dados_Atualizados.csv",
                            mime="text/csv"
                        )
                else:
                    st.error("Erro ao combinar os dados.")
    else:
        st.info("Por favor, primeiro carregue os dados de câmbio e o arquivo de destino nas abas anteriores.")

# Instruções de uso
st.sidebar.header("Instruções de Uso")
st.sidebar.markdown("""
### Passo a passo:

1. **Arquivo de Origem**:
   - Selecione o arquivo "Operações de câmbio BRA.xlsm"
   - Escolha o intervalo de datas
   - Clique em "Carregar Dados de Câmbio"
   - Verifique os dados extraídos

2. **Arquivo de Destino**:
   - Selecione o arquivo "01. Operações.xlsm"
   - Clique em "Carregar Arquivo de Destino"
   - Visualize os dados atuais

3. **Combinação**:
   - Vá para a aba "Arquivo Atualizado"
   - Clique em "Combinar Dados"
   - Visualize o resultado
   - Baixe o arquivo atualizado no formato desejado
""")

st.sidebar.markdown("---")
st.sidebar.info("""
**Observações importantes:**
- Os dados são adicionados no final da tabela existente
- A Data é inserida na coluna A
- A Receita BGX é inserida na coluna E
- O Cliente é inserido na coluna I
- Os formatos mais simples (XLSX e CSV) têm maior compatibilidade
""")
